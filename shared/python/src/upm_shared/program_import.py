"""Shared, persistence-free parsing for Central and Site program imports.

Database reconciliation deliberately remains in the owning deployment.  This module is the
single source of truth for accepted CSV/XLSX formats, column vocabulary, row normalization,
and source-local schedule parsing.
"""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime, time
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

MAX_IMPORT_BYTES = 25 * 1024 * 1024


COLUMN_ALIASES = {
    "presentation_date_added": "presentation_date_added",
    "presentation_id": "external_presentation_id",
    "presentationid": "external_presentation_id",
    "external_presentation_id": "external_presentation_id",
    "session_id": "session_code",
    "sessionid": "session_code",
    "session_code": "session_code",
    "presentation_title": "presentation_title",
    "session_title": "session_title",
    "session_name": "session_title",
    "presenterid": "external_id",
    "presenter_id": "external_id",
    "speakerid": "external_id",
    "speaker_id": "external_id",
    "personid": "external_id",
    "person_id": "external_id",
    "first_name": "given_name",
    "firstname": "given_name",
    "last_name": "family_name",
    "lastname": "family_name",
    "given_name": "given_name",
    "givenname": "given_name",
    "family_name": "family_name",
    "familyname": "family_name",
    "surname": "family_name",
    "company": "organization",
    "company_name": "organization",
    "employer": "organization",
    "organization": "organization",
    "position_title": "professional_title",
    "position": "professional_title",
    "job_title": "professional_title",
    "title_position": "professional_title",
    "presenter_role": "presenter_role",
    "speaker_role": "presenter_role",
    "role": "presenter_role",
    "presenter_roster_order": "presenter_order",
    "roster_order": "presenter_order",
    "presenter_order": "presenter_order",
    "speaker_order": "presenter_order",
    "display_order": "presenter_order",
    "type": "entity_type",
    "session": "session_code",
    "presentation": "presentation_code",
    "presentation_code": "presentation_code",
    "start": "starts_at",
    "start_time": "start_time",
    "session_start": "starts_at",
    "presentation_start": "starts_at",
    "starttime": "start_time",
    "end": "ends_at",
    "end_time": "end_time",
    "session_end": "ends_at",
    "presentation_end": "ends_at",
    "endtime": "end_time",
    "date": "session_date",
    "session_date": "session_date",
    "presentation_date": "session_date",
    "event_date": "session_date",
    "room": "location_name",
    "room_name": "location_name",
    "location": "location_name",
    "location_name": "location_name",
    "venue_room": "location_name",
    "speaker": "display_name",
    "speaker_name": "display_name",
    "presenter": "display_name",
    "presenter_name": "display_name",
    "display_name": "display_name",
    "speaker_email": "presenter_email",
    "presenter_email": "presenter_email",
    "email_address": "email",
    "e_mail": "email",
    "email": "email",
    "track": "track",
    "session_track": "track",
    "program_track": "track",
    "session_format": "session_format",
    "presentation_format": "session_format",
    "format": "session_format",
    "session_type": "session_format",
    "presentation_type": "session_format",
}


def normalize_text(value: str | None) -> str:
    return " ".join((value or "").strip().casefold().split())


def header_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.strip().casefold()).strip("_")


def _cell(value: object) -> object:
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def parse_csv(content: bytes) -> list[dict[str, object]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV must be UTF-8") from exc
    return [
        {str(key).strip(): _cell(value) for key, value in row.items()}
        for row in csv.DictReader(io.StringIO(text))
        if any(value not in (None, "") for value in row.values())
    ]


def parse_xlsx(content: bytes) -> list[dict[str, object]]:
    try:
        workbook = load_workbook(
            io.BytesIO(content), read_only=True, data_only=True, keep_links=False
        )
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(values)]
        rows: list[dict[str, object]] = []
        for row in values:
            raw = {
                header: _cell(value) for header, value in zip(headers, row, strict=False) if header
            }
            if any(value not in (None, "") for value in raw.values()):
                raw["__worksheet"] = sheet.title
                rows.append(raw)
        workbook.close()
        return rows
    except (
        BadZipFile,
        InvalidFileException,
        KeyError,
        OSError,
        TypeError,
        ValueError,
        StopIteration,
    ) as exc:
        raise ValueError("invalid XLSX workbook") from exc


def parse_program_source(filename: str, content: bytes) -> tuple[str, list[dict[str, object]]]:
    if len(content) > MAX_IMPORT_BYTES:
        raise ValueError(f"program import exceeds {MAX_IMPORT_BYTES} bytes")
    suffix = filename.rsplit(".", 1)[-1].casefold() if "." in filename else ""
    if suffix == "csv":
        return "csv", parse_csv(content)
    if suffix == "xlsx":
        return "xlsx", parse_xlsx(content)
    raise ValueError("program import must be CSV or XLSX")


def normalize_row(raw: dict[str, object]) -> dict[str, object]:
    result: dict[str, object] = {}
    for key, original in raw.items():
        if key == "__worksheet":
            result["_source_worksheet"] = original
            continue
        normalized_key = COLUMN_ALIASES.get(header_key(key), header_key(key))
        value = " ".join(original.strip().split()) if isinstance(original, str) else original
        if value in (None, "") and result.get(normalized_key) not in (None, ""):
            continue
        result[normalized_key] = value
    if result.get("presenter_email") and not result.get("email"):
        result["email"] = result["presenter_email"]
    if result.get("email"):
        result["normalized_email"] = normalize_text(str(result["email"]))
    display = (
        result.get("display_name")
        or " ".join(str(result.get(key) or "") for key in ("given_name", "family_name")).strip()
    )
    if display:
        result["display_name"] = display
        result["normalized_name"] = normalize_text(str(display))
    if result.get("external_id") and not result.get("external_namespace"):
        result["external_namespace"] = "program_import_presenter"
    return result


def detect_columns(headers: list[str]) -> dict[str, str]:
    return {
        header: COLUMN_ALIASES.get(header_key(header), header_key(header))
        for header in headers
        if header.strip() and header != "__worksheet"
    }


def parse_source_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    supplied = str(value).strip()
    if "T" in supplied or " " in supplied:
        try:
            return datetime.fromisoformat(supplied).date()
        except ValueError:
            pass
    try:
        return date.fromisoformat(supplied)
    except ValueError:
        pass
    for pattern in ("%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(supplied, pattern).date()
        except ValueError:
            continue
    raise ValueError(f"invalid source date {supplied!r}")


def parse_source_time(value: object) -> time:
    if isinstance(value, datetime):
        return value.time().replace(tzinfo=None)
    if isinstance(value, time):
        return value.replace(tzinfo=None)
    supplied = str(value).strip()
    if "T" in supplied or " " in supplied and not supplied.upper().endswith(("AM", "PM")):
        try:
            return datetime.fromisoformat(supplied).time().replace(tzinfo=None)
        except ValueError:
            pass
    try:
        return time.fromisoformat(supplied).replace(tzinfo=None)
    except ValueError:
        pass
    for pattern in ("%I:%M %p", "%I:%M:%S %p"):
        try:
            return datetime.strptime(supplied.upper(), pattern).time()
        except ValueError:
            continue
    raise ValueError(f"invalid source time {supplied!r}")
