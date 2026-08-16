"""Auditable CSV/XLSX staging, identity reconciliation, and transactional commit."""

import csv
import hashlib
import io
import re
from datetime import UTC, datetime
from decimal import Decimal
from uuid import UUID
from zipfile import BadZipFile
from zoneinfo import ZoneInfo

from fastapi import HTTPException, status
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, selectinload

from upm_central.persistence.models import (
    Event,
    EventParticipation,
    ExternalIdentifier,
    ImportBatch,
    ImportRow,
    ImportSource,
    ImportValidationIssue,
    Person,
    Presentation,
    PresentationPresenter,
    PresentationSession,
    ReconciliationDecision,
    SessionParticipant,
    utc_now,
)
from upm_central.persistence.models import (
    Session as ProgramSession,
)
from upm_central.program import (
    audit,
    external_identifier,
    normalize_email,
    normalize_text,
    touch_event_program,
)
from upm_shared.enums import (
    ExternalEntityType,
    IdentityMatchOutcome,
    ImportEntityType,
    ImportProposedAction,
    ImportSourceType,
    ImportStatus,
    ImportValidationState,
    ParticipantStatus,
    PresentationWorkflowStatus,
    ReconciliationAction,
    SessionStatus,
    ValidationSeverity,
)

MAX_IMPORT_BYTES = 25 * 1024 * 1024


def _header_key(value: str) -> str:
    """Normalize source headings without losing the original heading in row evidence."""
    return re.sub(r"[^a-z0-9]+", "_", value.strip().casefold()).strip("_")


# This is the single importer vocabulary.  More-specific headings intentionally have
# their own entries; notably professional title and date-added can never fall through
# to the broad program title/date concepts.
COLUMN_ALIASES = {
    "presentation_date_added": "presentation_date_added",
    "presentation_id": "session_code",
    "presentationid": "session_code",
    "session_id": "session_code",
    "sessionid": "session_code",
    "presentation_title": "session_title",
    "session_title": "session_title",
    "session_name": "session_title",
    "presenterid": "external_id",
    "presenter_id": "external_id",
    "speakerid": "external_id",
    "speaker_id": "external_id",
    "personid": "external_id",
    "person_id": "external_id",
    "first_name": "given_name",
    "firstname": "given_name",
    "last_name": "family_name",
    "lastname": "family_name",
    "given_name": "given_name",
    "givenname": "given_name",
    "family_name": "family_name",
    "familyname": "family_name",
    "surname": "family_name",
    "company": "organization",
    "company_name": "organization",
    "employer": "organization",
    "position_title": "professional_title",
    "position": "professional_title",
    "job_title": "professional_title",
    "title_position": "professional_title",
    "presenter_role": "presenter_role",
    "speaker_role": "presenter_role",
    "role": "presenter_role",
    "presenter_roster_order": "presenter_order",
    "roster_order": "presenter_order",
    "presenter_order": "presenter_order",
    "speaker_order": "presenter_order",
    "display_order": "presenter_order",
    "type": "entity_type",
    "session": "session_code",
    "presentation": "presentation_code",
    "start": "starts_at",
    "start_time": "starts_at",
    "session_start": "starts_at",
    "presentation_start": "starts_at",
    "starttime": "starts_at",
    "end": "ends_at",
    "end_time": "ends_at",
    "session_end": "ends_at",
    "presentation_end": "ends_at",
    "endtime": "ends_at",
    "date": "session_date",
    "session_date": "session_date",
    "presentation_date": "session_date",
    "event_date": "session_date",
    "room": "location_name",
    "room_name": "location_name",
    "location": "location_name",
    "location_name": "location_name",
    "venue_room": "location_name",
    "speaker": "display_name",
    "speaker_name": "display_name",
    "presenter": "display_name",
    "presenter_name": "display_name",
    "speaker_email": "presenter_email",
    "presenter_email": "presenter_email",
    "email_address": "email",
    "e_mail": "email",
    "track": "track",
    "session_track": "track",
    "program_track": "track",
    "session_format": "session_format",
    "presentation_format": "session_format",
    "format": "session_format",
    "session_type": "session_format",
    "presentation_type": "session_format",
}


def _cell(value: object) -> object:
    if isinstance(value, datetime):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _parse_csv(content: bytes) -> list[dict[str, object]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY, detail="CSV must be UTF-8"
        ) from exc
    return [
        {str(k).strip(): _cell(v) for k, v in row.items()}
        for row in csv.DictReader(io.StringIO(text))
    ]


def _parse_xlsx(content: bytes) -> list[dict[str, object]]:
    try:
        workbook = load_workbook(
            io.BytesIO(content), read_only=True, data_only=True, keep_links=False
        )
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(values)]
        rows = []
        for row in values:
            raw = {
                header: _cell(value) for header, value in zip(headers, row, strict=False) if header
            }
            if any(value not in (None, "") for value in raw.values()):
                raw["__worksheet"] = sheet.title
                rows.append(raw)
        workbook.close()
        return rows
    except (
        BadZipFile,
        InvalidFileException,
        KeyError,
        OSError,
        TypeError,
        ValueError,
        StopIteration,
    ) as exc:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY, detail="invalid XLSX workbook"
        ) from exc


def _normalized(raw: dict[str, object]) -> dict[str, object]:
    result: dict[str, object] = {}
    for key, value in raw.items():
        if key == "__worksheet":
            result["_source_worksheet"] = value
            continue
        key_value = _header_key(key)
        normalized_key = COLUMN_ALIASES.get(key_value, key_value)
        if isinstance(value, str):
            value = " ".join(value.strip().split())
        result[normalized_key] = value
    if result.get("presenter_email") and not result.get("email"):
        result["email"] = result["presenter_email"]
    if result.get("email"):
        result["normalized_email"] = normalize_text(str(result["email"]))
    display = (
        result.get("display_name")
        or " ".join(str(result.get(key) or "") for key in ("given_name", "family_name")).strip()
    )
    if display:
        result["display_name"] = display
        result["normalized_name"] = normalize_text(str(display))
    if result.get("external_id") and not result.get("external_namespace"):
        result["external_namespace"] = "program_import_presenter"
    date = result.get("session_date")
    if date:
        for source, target in (("starts_at", "starts_at"), ("ends_at", "ends_at")):
            if result.get(source) and "T" not in str(result[source]):
                result[target] = f"{date}T{result[source]}"
    return result


def detect_columns(headers: list[str]) -> dict[str, str]:
    """Return the importer field selected by the existing normalization rules."""
    result = {}
    for header in headers:
        if not header.strip() or header == "__worksheet":
            continue
        normalized = _header_key(header)
        result[header] = COLUMN_ALIASES.get(normalized, normalized)
    return result


def _entity(values: dict[str, object]) -> ImportEntityType:
    supplied = normalize_text(str(values.get("entity_type") or ""))
    mapping = {
        "person": ImportEntityType.PERSON,
        "participant": ImportEntityType.PARTICIPANT,
        "presenter": ImportEntityType.PARTICIPANT,
        "session": ImportEntityType.SESSION,
        "presentation": ImportEntityType.PRESENTATION,
        "relationship": ImportEntityType.RELATIONSHIP,
    }
    if supplied in mapping:
        return mapping[supplied]
    if values.get("presentation_code") or values.get("presentation_title"):
        return ImportEntityType.PRESENTATION
    if values.get("session_code") or values.get("session_title"):
        return ImportEntityType.SESSION
    if values.get("email") or values.get("display_name"):
        return ImportEntityType.PARTICIPANT
    return ImportEntityType.UNKNOWN


def _schedule(values: dict[str, object], event: Event) -> tuple[dict[str, object], list[str]]:
    result = dict(values)
    errors: list[str] = []
    parsed: dict[str, datetime] = {}
    for field in ("starts_at", "ends_at", "scheduled_at"):
        raw = values.get(field)
        if not raw:
            continue
        try:
            value = datetime.fromisoformat(str(raw).replace("Z", "+00:00"))
            if value.utcoffset() is None:
                zone = ZoneInfo(event.timezone)
                first = value.replace(tzinfo=zone, fold=0)
                second = value.replace(tzinfo=zone, fold=1)
                if first.utcoffset() != second.utcoffset():
                    raise ValueError(
                        "ambiguous or nonexistent local time requires an explicit offset"
                    )
                value = first
            value = value.astimezone(UTC)
            parsed[field] = value
            result[field] = value.isoformat()
        except ValueError as exc:
            errors.append(f"{field}: {exc}")
    if parsed.get("starts_at") and parsed.get("ends_at"):
        if parsed["ends_at"] <= parsed["starts_at"]:
            errors.append("ends_at must follow starts_at")
    return result, errors


def _issue(
    row: ImportRow, severity: ValidationSeverity, code: str, message: str, field: str | None = None
) -> ImportValidationIssue:
    return ImportValidationIssue(
        row=row, severity=severity, code=code, field_name=field, message=message, details={}
    )


def _match_person(session: Session, row: ImportRow, values: dict[str, object]) -> None:
    supplied_id = values.get("person_id")
    if supplied_id:
        try:
            person_id = UUID(str(supplied_id))
        except ValueError:
            row.match_outcome = IdentityMatchOutcome.CONFLICT
            row.match_reason = "supplied person_id is malformed"
            return
        person = session.get(Person, person_id)
        if person is None or person.deleted_at is not None:
            row.match_outcome = IdentityMatchOutcome.CONFLICT
            row.match_reason = "supplied person_id does not identify an active person"
            return
        row.match_outcome = IdentityMatchOutcome.EXACT
        row.proposed_person_id = person.person_id
        row.candidate_person_ids = [str(person.person_id)]
        row.match_confidence = Decimal("1.0000")
        row.match_reason = "authoritative UPM person_id"
        row.proposed_action = ImportProposedAction.MATCH_EXISTING
        row.normalized_values = {
            **row.normalized_values,
            "_matched_person_revision": person.revision,
        }
        return
    namespace, value = values.get("external_namespace"), values.get("external_id")
    if namespace and value:
        identifier = session.scalar(
            select(ExternalIdentifier).where(
                ExternalIdentifier.namespace == normalize_text(str(namespace)),
                ExternalIdentifier.normalized_external_id == normalize_text(str(value)),
            )
        )
        if identifier and identifier.entity_type == ExternalEntityType.PERSON:
            person = session.get(Person, identifier.entity_id)
            if person and person.deleted_at is None:
                row.match_outcome = IdentityMatchOutcome.EXACT
                row.proposed_person_id = person.person_id
                row.candidate_person_ids = [str(person.person_id)]
                row.match_confidence = Decimal("1.0000")
                row.match_reason = "unique external identity"
                row.proposed_action = ImportProposedAction.MATCH_EXISTING
                row.normalized_values = {
                    **row.normalized_values,
                    "_matched_person_revision": person.revision,
                }
                return
    email = normalize_email(str(values.get("email") or ""))
    if email:
        people = session.scalars(
            select(Person).where(Person.normalized_email == email, Person.deleted_at.is_(None))
        ).all()
        if len(people) == 1:
            person = people[0]
            row.match_outcome = IdentityMatchOutcome.EXACT
            row.proposed_person_id = person.person_id
            row.candidate_person_ids = [str(person.person_id)]
            row.match_confidence = Decimal("0.9500")
            row.match_reason = "unique normalized email"
            row.proposed_action = ImportProposedAction.MATCH_EXISTING
            row.normalized_values = {
                **row.normalized_values,
                "_matched_person_revision": person.revision,
            }
            return
        if len(people) > 1:
            row.match_outcome = IdentityMatchOutcome.AMBIGUOUS
            row.candidate_person_ids = [str(person.person_id) for person in people]
            row.match_confidence = Decimal("0.5000")
            row.match_reason = "normalized email maps to multiple people"
            row.conflict_state = "operator_review_required"
            return
    name = values.get("normalized_name")
    if name:
        people = session.scalars(
            select(Person).where(Person.normalized_name == name, Person.deleted_at.is_(None))
        ).all()
        if people:
            row.match_outcome = (
                IdentityMatchOutcome.STRONG_CANDIDATE
                if len(people) == 1
                else IdentityMatchOutcome.AMBIGUOUS
            )
            row.candidate_person_ids = [str(person.person_id) for person in people]
            row.match_confidence = Decimal("0.4000")
            row.match_reason = "name-only evidence is never auto-merged"
            row.conflict_state = "operator_review_required"
            return
    row.match_outcome = IdentityMatchOutcome.NO_MATCH
    row.match_confidence = Decimal("0.0000")
    row.match_reason = "no authoritative identity evidence"
    row.proposed_action = ImportProposedAction.CREATE_NEW


def create_batch(
    session: Session,
    *,
    event: Event,
    filename: str,
    content_type: str,
    content: bytes,
    actor: str = "central-admin",
) -> ImportBatch:
    if not content or len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(
            status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="import size is invalid"
        )
    lower = filename.casefold()
    if lower.endswith(".csv"):
        source_type, parsed = ImportSourceType.CSV, _parse_csv(content)
    elif lower.endswith(".xlsx"):
        source_type, parsed = ImportSourceType.XLSX, _parse_xlsx(content)
    else:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY, detail="only CSV and XLSX are supported"
        )
    digest = hashlib.sha256(content).hexdigest()
    duplicate = session.scalar(
        select(ImportBatch).where(
            ImportBatch.event_id == event.event_id,
            ImportBatch.source_sha256 == digest,
            ImportBatch.importer_type == "program",
        )
    )
    if duplicate is not None:
        return duplicate
    source = ImportSource(
        filename=filename,
        content_type=content_type,
        source_type=source_type,
        size_bytes=len(content),
        sha256=digest,
        content=content,
        uploaded_by=actor,
    )
    session.add(source)
    session.flush()
    batch = ImportBatch(
        event_id=event.event_id,
        import_source_id=source.import_source_id,
        filename=filename,
        source_sha256=digest,
        importer_type="program",
        status=ImportStatus.PARSING,
        created_by=actor,
        reviewed_domain_revision=event.revision,
    )
    session.add(batch)
    session.flush()
    grouped_sessions: dict[str, tuple[tuple[object, ...], ImportRow]] = {}
    seen_presentation_codes: set[str] = set()
    normalized_rows = [_normalized(raw) for raw in parsed]
    imported_emails = {
        email
        for values in normalized_rows
        if (email := normalize_email(str(values.get("email") or "")))
    }
    imported_session_codes = {
        str(values["session_code"])
        for values in normalized_rows
        if _entity(values) == ImportEntityType.SESSION and values.get("session_code")
    }
    known_event_emails = set(
        session.scalars(
            select(Person.normalized_email)
            .join(EventParticipation, EventParticipation.person_id == Person.person_id)
            .where(
                EventParticipation.event_id == event.event_id,
                Person.normalized_email.is_not(None),
            )
        )
    )
    known_session_codes = set(
        session.scalars(
            select(ProgramSession.session_code).where(
                ProgramSession.event_id == event.event_id,
                ProgramSession.session_code.is_not(None),
            )
        )
    )
    for number, (raw, values) in enumerate(zip(parsed, normalized_rows, strict=True), start=2):
        values, schedule_errors = _schedule(values, event)
        row = ImportRow(
            import_batch_id=batch.import_batch_id,
            source_row_number=number,
            raw_values=raw,
            normalized_values=values,
            entity_type=_entity(values),
            validation_state=ImportValidationState.PENDING,
        )
        session.add(row)
        session.flush()
        for message in schedule_errors:
            row.issues.append(
                _issue(
                    row,
                    ValidationSeverity.ERROR,
                    "invalid_schedule",
                    message,
                )
            )
        code_sets = []
        # Repeated session codes are roster evidence, not duplicate entities.  Only
        # disagreements in authoritative program fields are blocking conflicts.
        if row.entity_type == ImportEntityType.PRESENTATION:
            code_sets.append(
                ("presentation_code", seen_presentation_codes, "duplicate_presentation_code")
            )
        for field, seen, issue_code in code_sets:
            value = normalize_text(str(values.get(field) or ""))
            if value and value in seen:
                row.issues.append(
                    _issue(row, ValidationSeverity.ERROR, issue_code, f"Duplicate {field}", field)
                )
            elif value:
                seen.add(value)
        # Repeated person identifiers are expected roster evidence.  The identity
        # resolver reuses their permanent Person rather than rejecting later rows.
        if row.entity_type in {ImportEntityType.PERSON, ImportEntityType.PARTICIPANT}:
            if not values.get("display_name"):
                row.issues.append(
                    _issue(
                        row,
                        ValidationSeverity.ERROR,
                        "missing_name",
                        "A person or participant requires a name",
                        "display_name",
                    )
                )
            if values.get("email") and normalize_email(str(values["email"])) is None:
                row.issues.append(
                    _issue(
                        row,
                        ValidationSeverity.ERROR,
                        "invalid_email",
                        "Email address is malformed",
                        "email",
                    )
                )
            _match_person(session, row, values)
            if row.match_outcome in {
                IdentityMatchOutcome.AMBIGUOUS,
                IdentityMatchOutcome.STRONG_CANDIDATE,
                IdentityMatchOutcome.CONFLICT,
            }:
                row.issues.append(
                    _issue(
                        row,
                        ValidationSeverity.ERROR,
                        "identity_review",
                        row.match_reason or "Identity requires review",
                    )
                )
        elif row.entity_type == ImportEntityType.SESSION:
            if not (values.get("title") or values.get("session_title")):
                row.issues.append(
                    _issue(
                        row,
                        ValidationSeverity.ERROR,
                        "missing_session_title",
                        "Session title is required",
                        "title",
                    )
                )
            row.proposed_action = ImportProposedAction.CREATE_OR_UPDATE
        elif row.entity_type == ImportEntityType.PRESENTATION:
            if not (
                values.get("title")
                or values.get("presentation_title")
                or values.get("session_title")
            ):
                row.issues.append(
                    _issue(
                        row,
                        ValidationSeverity.ERROR,
                        "missing_presentation_title",
                        "Presentation title is required",
                        "title",
                    )
                )
            row.proposed_action = ImportProposedAction.CREATE_OR_UPDATE
        else:
            row.issues.append(
                _issue(
                    row,
                    ValidationSeverity.ERROR,
                    "unknown_entity",
                    "The row entity type could not be detected",
                )
            )
        if row.entity_type in {ImportEntityType.SESSION, ImportEntityType.PRESENTATION}:
            if _presenter_emails(values) and values.get("display_name"):
                _match_person(session, row, values)
                if row.match_outcome in {
                    IdentityMatchOutcome.AMBIGUOUS,
                    IdentityMatchOutcome.STRONG_CANDIDATE,
                    IdentityMatchOutcome.CONFLICT,
                }:
                    row.issues.append(
                        _issue(
                            row,
                            ValidationSeverity.ERROR,
                            "identity_review",
                            row.match_reason or "Presenter identity requires review",
                        )
                    )
        if row.entity_type == ImportEntityType.SESSION and values.get("session_code"):
            group_key = normalize_text(str(values["session_code"])) or ""
            signature = tuple(
                values.get(field)
                for field in (
                    "session_title",
                    "session_date",
                    "starts_at",
                    "ends_at",
                    "location_name",
                    "track",
                    "session_format",
                )
            )
            previous = grouped_sessions.get(group_key)
            if previous and previous[0] != signature:
                message = "Rows with this program identifier disagree on authoritative session data"
                row.issues.append(
                    _issue(
                        row,
                        ValidationSeverity.ERROR,
                        "grouped_session_conflict",
                        message,
                        "session_code",
                    )
                )
                previous[1].issues.append(
                    _issue(
                        previous[1],
                        ValidationSeverity.ERROR,
                        "grouped_session_conflict",
                        message,
                        "session_code",
                    )
                )
                row.conflict_state = previous[1].conflict_state = "operator_review_required"
            elif previous is None:
                grouped_sessions[group_key] = (signature, row)
            for email in _presenter_emails(values):
                if email not in imported_emails and email not in known_event_emails:
                    row.issues.append(
                        _issue(
                            row,
                            ValidationSeverity.ERROR,
                            "unresolved_presenter",
                            f"Presenter email {email} is not an event participant",
                            "presenter_email",
                        )
                    )
        if row.entity_type == ImportEntityType.PRESENTATION and values.get("session_code"):
            code = str(values["session_code"])
            if (
                code not in imported_session_codes
                and code not in known_session_codes
                and not values.get("session_title")
            ):
                row.issues.append(
                    _issue(
                        row,
                        ValidationSeverity.ERROR,
                        "unresolved_session",
                        f"Session code {code} cannot be resolved",
                        "session_code",
                    )
                )
        row.validation_state = (
            ImportValidationState.ERROR
            if any(i.severity == ValidationSeverity.ERROR for i in row.issues)
            else ImportValidationState.WARNING
            if row.issues
            else ImportValidationState.VALID
        )
    session.flush()
    rows = session.scalars(
        select(ImportRow).where(ImportRow.import_batch_id == batch.import_batch_id)
    ).all()
    for row in rows:
        row.validation_state = (
            ImportValidationState.ERROR
            if any(issue.severity == ValidationSeverity.ERROR for issue in row.issues)
            else ImportValidationState.WARNING
            if row.issues
            else ImportValidationState.VALID
        )
    batch.row_count = len(rows)
    batch.valid_count = sum(row.validation_state == ImportValidationState.VALID for row in rows)
    batch.warning_count = sum(row.validation_state == ImportValidationState.WARNING for row in rows)
    batch.conflict_count = sum(bool(row.conflict_state) for row in rows)
    batch.status = (
        ImportStatus.REVIEW
        if any(row.validation_state == ImportValidationState.ERROR for row in rows)
        else ImportStatus.READY
    )
    audit(
        session,
        action="central.import_batch.created",
        target_type="import_batch",
        target_id=batch.import_batch_id,
        event_id=event.event_id,
        after={"filename": filename, "sha256": digest, "rows": len(rows)},
        actor=actor,
    )
    return batch


def decide(
    session: Session,
    row: ImportRow,
    *,
    action: ReconciliationAction,
    selected_person_id: UUID | None,
    corrected_values: dict[str, object] | None,
    reason: str | None,
    actor: str = "central-admin",
) -> ReconciliationDecision:
    if row.committed_at is not None:
        raise HTTPException(status.HTTP_409_CONFLICT, detail="row is already committed")
    had_conflict = bool(row.conflict_state)
    if action in {ReconciliationAction.ACCEPT_MATCH, ReconciliationAction.CHOOSE_PERSON}:
        selected_person_id = selected_person_id or row.proposed_person_id
        person = session.get(Person, selected_person_id) if selected_person_id else None
        if person is None or person.deleted_at is not None:
            raise HTTPException(
                status.HTTP_422_UNPROCESSABLE_ENTITY, detail="active person required"
            )
        row.resolved_person_id = person.person_id
        row.normalized_values = {
            **row.normalized_values,
            "_matched_person_revision": person.revision,
        }
    elif action == ReconciliationAction.CREATE_PERSON:
        row.resolved_person_id = None
    row.resolution_action = action
    row.corrected_values = corrected_values
    row.resolved_by = actor
    row.resolved_at = utc_now()
    row.conflict_state = None
    row.validation_state = ImportValidationState.VALID
    if had_conflict:
        row.batch.conflict_count = max(0, row.batch.conflict_count - 1)
    if row.batch.conflict_count == 0:
        row.batch.status = ImportStatus.READY
    decision = ReconciliationDecision(
        import_row_id=row.import_row_id,
        action=action,
        selected_person_id=selected_person_id,
        corrected_values=corrected_values,
        decided_by=actor,
        reason=reason,
    )
    session.add(decision)
    audit(
        session,
        action="central.import_row.reconciled",
        target_type="import_row",
        target_id=row.import_row_id,
        event_id=row.batch.event_id,
        after={
            "action": action,
            "selected_person_id": str(selected_person_id) if selected_person_id else None,
        },
        actor=actor,
    )
    return decision


def _person_for_row(
    session: Session, row: ImportRow, values: dict[str, object], actor: str
) -> Person | None:
    if row.resolution_action in {ReconciliationAction.IGNORE, ReconciliationAction.REJECT}:
        return None
    person_id = row.resolved_person_id or row.proposed_person_id
    if person_id:
        person = session.get(Person, person_id)
        if person is None:
            raise HTTPException(status.HTTP_409_CONFLICT, detail="matched person no longer exists")
        reviewed_revision = values.get("_matched_person_revision")
        if reviewed_revision is not None and person.revision != int(reviewed_revision):
            raise HTTPException(
                status.HTTP_409_CONFLICT, detail="matched person changed; revalidate import"
            )
        return person
    display_name = str(values.get("display_name") or "").strip()
    person = Person(
        given_name=values.get("given_name") or None,
        family_name=values.get("family_name") or None,
        display_name=display_name,
        normalized_name=normalize_text(display_name) or "",
        primary_email=values.get("email") or None,
        normalized_email=normalize_email(str(values.get("email") or "")),
        organization=values.get("organization") or None,
        professional_title=values.get("professional_title") or None,
    )
    session.add(person)
    session.flush()
    audit(
        session,
        action="central.person.created",
        target_type="person",
        target_id=person.person_id,
        after={"display_name": display_name},
        actor=actor,
    )
    return person


def _presenter_emails(values: dict[str, object]) -> list[str]:
    supplied = (
        values.get("presenter_emails") or values.get("presenter_email") or values.get("email")
    )
    if not supplied:
        return []
    return [
        email
        for item in str(supplied).replace(";", ",").split(",")
        if (email := normalize_email(item))
    ]


def _participant_by_email(
    session: Session, event_id: UUID, email: str
) -> EventParticipation | None:
    return session.scalar(
        select(EventParticipation)
        .join(Person, Person.person_id == EventParticipation.person_id)
        .where(
            EventParticipation.event_id == event_id,
            Person.normalized_email == email,
        )
    )


def _ensure_wide_row_participant(
    session: Session,
    event: Event,
    row: ImportRow,
    values: dict[str, object],
    email: str,
    actor: str,
) -> EventParticipation:
    participant = _participant_by_email(session, event.event_id, email)
    if participant is not None:
        participant.is_presenter = True
        return participant
    person_id = row.resolved_person_id or row.proposed_person_id
    person = session.get(Person, person_id) if person_id else None
    if person is None:
        people = session.scalars(
            select(Person).where(Person.normalized_email == email, Person.deleted_at.is_(None))
        ).all()
        if len(people) > 1:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                detail=f"presenter email {email} is ambiguous; reconcile before commit",
            )
        person = people[0] if people else None
    if person is None:
        display_name = str(values.get("display_name") or email).strip()
        person = Person(
            given_name=values.get("given_name") or None,
            family_name=values.get("family_name") or None,
            display_name=display_name,
            normalized_name=normalize_text(display_name) or "",
            primary_email=email,
            normalized_email=email,
            organization=values.get("organization") or None,
            professional_title=values.get("professional_title") or None,
        )
        session.add(person)
        session.flush()
        audit(
            session,
            action="central.person.created",
            target_type="person",
            target_id=person.person_id,
            after={"display_name": display_name},
            actor=actor,
        )
    if values.get("external_id"):
        external_identifier(
            session,
            entity_type=ExternalEntityType.PERSON,
            entity_id=person.person_id,
            namespace=str(values.get("external_namespace") or "program_import_presenter"),
            value=str(values["external_id"]),
            event_id=None,
            source="import",
        )
    participant = EventParticipation(
        event_id=event.event_id,
        person_id=person.person_id,
        professional_title=values.get("professional_title") or None,
        organization=values.get("organization") or None,
        participant_status=ParticipantStatus.ACTIVE,
        is_presenter=True,
        source="import",
        source_metadata={"import_batch_id": str(row.import_batch_id)},
    )
    session.add(participant)
    session.flush()
    return participant


def commit_batch(
    session: Session, batch: ImportBatch, *, actor: str = "central-admin"
) -> ImportBatch:
    if batch.status == ImportStatus.COMMITTED:
        return batch
    if batch.status not in {ImportStatus.READY, ImportStatus.REVIEW}:
        raise HTTPException(status.HTTP_409_CONFLICT, detail="import is not reviewable")
    event = session.get(Event, batch.event_id)
    if event is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="event not found")
    if batch.reviewed_domain_revision != event.revision:
        raise HTTPException(
            status.HTTP_409_CONFLICT, detail="event program changed; revalidate import"
        )
    rows = session.scalars(
        select(ImportRow)
        .where(ImportRow.import_batch_id == batch.import_batch_id)
        .options(selectinload(ImportRow.issues))
        .order_by(ImportRow.source_row_number)
    ).all()
    unresolved = [
        row
        for row in rows
        if row.validation_state == ImportValidationState.ERROR
        and row.resolution_action not in {ReconciliationAction.IGNORE, ReconciliationAction.REJECT}
    ]
    if unresolved:
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            detail={
                "message": "blocking import conflicts remain",
                "rows": [r.source_row_number for r in unresolved],
            },
        )
    batch.status = ImportStatus.COMMITTING
    changed = False
    priority = {
        ImportEntityType.PERSON: 0,
        ImportEntityType.PARTICIPANT: 0,
        ImportEntityType.SESSION: 1,
        ImportEntityType.PRESENTATION: 2,
        ImportEntityType.RELATIONSHIP: 3,
        ImportEntityType.UNKNOWN: 4,
    }
    rows.sort(key=lambda row: (priority[row.entity_type], row.source_row_number))
    try:
        for row in rows:
            if row.committed_at is not None:
                continue
            values = dict(row.normalized_values)
            if row.corrected_values:
                values.update(row.corrected_values)
            if row.resolution_action in {ReconciliationAction.IGNORE, ReconciliationAction.REJECT}:
                batch.rejected_count += 1
                row.committed_at = utc_now()
                audit(
                    session,
                    action="central.import_row.rejected",
                    target_type="import_row",
                    target_id=row.import_row_id,
                    event_id=event.event_id,
                    after={"resolution_action": row.resolution_action},
                    actor=actor,
                )
                continue
            if row.entity_type in {ImportEntityType.PERSON, ImportEntityType.PARTICIPANT}:
                person = _person_for_row(session, row, values, actor)
                if person is None:
                    continue
                ids: dict[str, object] = {"person_id": str(person.person_id)}
                if row.entity_type == ImportEntityType.PARTICIPANT:
                    participant = session.scalar(
                        select(EventParticipation).where(
                            EventParticipation.event_id == event.event_id,
                            EventParticipation.person_id == person.person_id,
                        )
                    )
                    if participant is None:
                        participant = EventParticipation(
                            event_id=event.event_id,
                            person_id=person.person_id,
                            display_name=values.get("event_display_name") or None,
                            professional_title=values.get("professional_title") or None,
                            organization=values.get("organization") or None,
                            participant_status=ParticipantStatus.ACTIVE,
                            is_presenter=str(values.get("is_presenter") or "").casefold()
                            in {"1", "true", "yes"},
                            source="import",
                            source_metadata={"import_batch_id": str(batch.import_batch_id)},
                        )
                        session.add(participant)
                        session.flush()
                        audit(
                            session,
                            action="central.event_participant.created",
                            target_type="event_participation",
                            target_id=participant.event_participation_id,
                            event_id=event.event_id,
                            after={"person_id": str(person.person_id)},
                            actor=actor,
                        )
                    ids["event_participation_id"] = str(participant.event_participation_id)
                if values.get("external_namespace") and values.get("external_id"):
                    external_identifier(
                        session,
                        entity_type=ExternalEntityType.PERSON,
                        entity_id=person.person_id,
                        namespace=str(values["external_namespace"]),
                        value=str(values["external_id"]),
                        event_id=None,
                        source="import",
                    )
                row.committed_entity_ids = ids
                changed = True
            elif row.entity_type == ImportEntityType.SESSION:
                title = str(values.get("title") or values.get("session_title"))
                code = values.get("session_code") or None
                item = (
                    session.scalar(
                        select(ProgramSession).where(
                            ProgramSession.event_id == event.event_id,
                            ProgramSession.session_code == code,
                        )
                    )
                    if code
                    else None
                )
                created = item is None
                if item is None:
                    item = ProgramSession(
                        event_id=event.event_id,
                        title=title,
                        session_code=code,
                        status=SessionStatus.SCHEDULED,
                        source="import",
                        source_metadata={"import_batch_id": str(batch.import_batch_id)},
                    )
                    session.add(item)
                else:
                    item.title = title
                    item.revision += 1
                item.starts_at = (
                    datetime.fromisoformat(str(values["starts_at"]))
                    if values.get("starts_at")
                    else None
                )
                item.ends_at = (
                    datetime.fromisoformat(str(values["ends_at"]))
                    if values.get("ends_at")
                    else None
                )
                imported_room = str(values.get("location_name") or "").strip() or None
                item.location_name = imported_room
                item.location_metadata = {
                    **(item.location_metadata or {}),
                    "imported_label": imported_room,
                    "normalized_imported_label": normalize_text(imported_room),
                    "mapping_state": "unmapped" if imported_room else "unassigned",
                    "import_batch_id": str(batch.import_batch_id),
                    "source_row_numbers": sorted(
                        set((item.location_metadata or {}).get("source_row_numbers", []))
                        | {row.source_row_number}
                    ),
                }
                item.session_type = str(values.get("session_format") or "").strip() or None
                item.source_metadata = {
                    **(item.source_metadata or {}),
                    "import_batch_id": str(batch.import_batch_id),
                    "program_external_id": code,
                    "track": values.get("track") or None,
                    "presentation_date_added": values.get("presentation_date_added") or None,
                    "source_row_ids": sorted(
                        set((item.source_metadata or {}).get("source_row_ids", []))
                        | {str(row.import_row_id)}
                    ),
                }
                session.flush()
                participant_ids: list[str] = []
                person_ids: list[str] = []
                relationship_ids: list[str] = []
                supplied_order = int(float(values.get("presenter_order") or 0))
                supplied_role = str(values.get("presenter_role") or "presenter").strip()
                for offset, email in enumerate(_presenter_emails(values)):
                    participant = _ensure_wide_row_participant(
                        session, event, row, values, email, actor
                    )
                    participant_ids.append(str(participant.event_participation_id))
                    person_ids.append(str(participant.person_id))
                    if (
                        session.scalar(
                            select(SessionParticipant).where(
                                SessionParticipant.session_id == item.session_id,
                                SessionParticipant.event_participation_id
                                == participant.event_participation_id,
                                SessionParticipant.role == supplied_role,
                            )
                        )
                        is None
                    ):
                        relationship = SessionParticipant(
                            session_id=item.session_id,
                            event_participation_id=participant.event_participation_id,
                            role=supplied_role,
                            presenter_order=supplied_order + offset,
                            primary_presenter=(supplied_order + offset) == 0,
                            external_relationship_id=str(row.import_row_id),
                            source="import",
                        )
                        session.add(relationship)
                        session.flush()
                    else:
                        relationship = session.scalar(
                            select(SessionParticipant).where(
                                SessionParticipant.session_id == item.session_id,
                                SessionParticipant.event_participation_id
                                == participant.event_participation_id,
                                SessionParticipant.role == supplied_role,
                            )
                        )
                    relationship_ids.append(str(relationship.session_participant_id))
                row.committed_entity_ids = {
                    "session_id": str(item.session_id),
                    "person_ids": person_ids,
                    "event_participation_ids": participant_ids,
                    "session_participant_ids": relationship_ids,
                }
                audit(
                    session,
                    action=f"central.session.{'created' if created else 'updated'}",
                    target_type="session",
                    target_id=item.session_id,
                    event_id=event.event_id,
                    after={"title": title},
                    actor=actor,
                )
                changed = True
            elif row.entity_type == ImportEntityType.PRESENTATION:
                title = str(
                    values.get("title")
                    or values.get("presentation_title")
                    or values.get("session_title")
                )
                code = values.get("presentation_code") or None
                item = (
                    session.scalar(
                        select(Presentation).where(
                            Presentation.event_id == event.event_id,
                            Presentation.presentation_code == code,
                        )
                    )
                    if code
                    else None
                )
                created = item is None
                if item is None:
                    item = Presentation(
                        event_id=event.event_id,
                        title=title,
                        presentation_code=code,
                        workflow_status=PresentationWorkflowStatus.EXPECTED,
                        source="import",
                        source_metadata={"import_batch_id": str(batch.import_batch_id)},
                    )
                    session.add(item)
                else:
                    item.title = title
                    item.revision += 1
                item.scheduled_at = (
                    datetime.fromisoformat(str(values["scheduled_at"]))
                    if values.get("scheduled_at")
                    else None
                )
                session.flush()
                session_code = values.get("session_code")
                target_session = None
                if session_code:
                    target_session = session.scalar(
                        select(ProgramSession).where(
                            ProgramSession.event_id == event.event_id,
                            ProgramSession.session_code == session_code,
                        )
                    )
                    if target_session is None and values.get("session_title"):
                        target_session = ProgramSession(
                            event_id=event.event_id,
                            title=str(values["session_title"]),
                            session_code=session_code,
                            starts_at=(
                                datetime.fromisoformat(str(values["starts_at"]))
                                if values.get("starts_at")
                                else None
                            ),
                            ends_at=(
                                datetime.fromisoformat(str(values["ends_at"]))
                                if values.get("ends_at")
                                else None
                            ),
                            location_name=str(values.get("location_name") or "").strip() or None,
                            location_metadata={
                                "imported_label": values.get("location_name") or None,
                                "normalized_imported_label": normalize_text(
                                    str(values.get("location_name") or "")
                                ),
                                "mapping_state": (
                                    "unmapped" if values.get("location_name") else "unassigned"
                                ),
                                "import_batch_id": str(batch.import_batch_id),
                            },
                            status=SessionStatus.SCHEDULED,
                            source="import",
                            source_metadata={"import_batch_id": str(batch.import_batch_id)},
                        )
                        session.add(target_session)
                        session.flush()
                    if target_session is None:
                        raise HTTPException(
                            status.HTTP_409_CONFLICT,
                            detail=f"unresolved session_code {session_code}",
                        )
                    item.session_id = target_session.session_id
                    if (
                        session.scalar(
                            select(PresentationSession).where(
                                PresentationSession.presentation_id == item.presentation_id,
                                PresentationSession.session_id == target_session.session_id,
                            )
                        )
                        is None
                    ):
                        session.add(
                            PresentationSession(
                                presentation_id=item.presentation_id,
                                session_id=target_session.session_id,
                                association_type="scheduled",
                                primary_session=True,
                                source="import",
                            )
                        )
                participant_ids = []
                person_ids = []
                for order, email in enumerate(_presenter_emails(values)):
                    participant = _ensure_wide_row_participant(
                        session, event, row, values, email, actor
                    )
                    participant_ids.append(str(participant.event_participation_id))
                    person_ids.append(str(participant.person_id))
                    if (
                        target_session
                        and session.scalar(
                            select(SessionParticipant).where(
                                SessionParticipant.session_id == target_session.session_id,
                                SessionParticipant.event_participation_id
                                == participant.event_participation_id,
                                SessionParticipant.role == "presenter",
                            )
                        )
                        is None
                    ):
                        session.add(
                            SessionParticipant(
                                session_id=target_session.session_id,
                                event_participation_id=participant.event_participation_id,
                                role="presenter",
                                presenter_order=order,
                                primary_presenter=order == 0,
                                source="import",
                            )
                        )
                    if (
                        session.scalar(
                            select(PresentationPresenter).where(
                                PresentationPresenter.presentation_id == item.presentation_id,
                                PresentationPresenter.event_participation_id
                                == participant.event_participation_id,
                                PresentationPresenter.role == "presenter",
                            )
                        )
                        is None
                    ):
                        session.add(
                            PresentationPresenter(
                                presentation_id=item.presentation_id,
                                event_participation_id=participant.event_participation_id,
                                role="presenter",
                                presenter_order=order,
                                primary_presenter=order == 0,
                                source="import",
                            )
                        )
                row.committed_entity_ids = {
                    "presentation_id": str(item.presentation_id),
                    "session_id": str(target_session.session_id) if target_session else None,
                    "person_ids": person_ids,
                    "event_participation_ids": participant_ids,
                }
                audit(
                    session,
                    action=f"central.presentation.{'created' if created else 'updated'}",
                    target_type="presentation",
                    target_id=item.presentation_id,
                    event_id=event.event_id,
                    after={"title": title},
                    actor=actor,
                )
                changed = True
            row.committed_at = utc_now()
            batch.committed_count += 1
            audit(
                session,
                action="central.import_row.committed",
                target_type="import_row",
                target_id=row.import_row_id,
                event_id=event.event_id,
                after={
                    "entity_type": row.entity_type,
                    "proposed_action": row.proposed_action,
                    "committed_entity_ids": row.committed_entity_ids,
                },
                actor=actor,
            )
        if changed:
            touch_event_program(session, event)
        batch.status = ImportStatus.COMMITTED
        batch.committed_at = utc_now()
        audit(
            session,
            action="central.import_batch.committed",
            target_type="import_batch",
            target_id=batch.import_batch_id,
            event_id=event.event_id,
            after={"committed": batch.committed_count, "rejected": batch.rejected_count},
            actor=actor,
        )
        return batch
    except IntegrityError as exc:
        raise HTTPException(
            status.HTTP_409_CONFLICT, detail="import conflicts with current domain data"
        ) from exc
